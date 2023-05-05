#!/usr/bin/env python
# _*_ coding:utf-8 _*_

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

# try to fix excel styler slow issue
# https://github.com/pandas-dev/pandas/issues/19917
from pandas.io.formats.style import Styler


def _update_ctx(self, attrs):
    rows = [(row_label, v) for row_label, v in attrs.iterrows()]
    row_idx = self.index.get_indexer([x[0] for x in rows])
    for ii, row in enumerate(rows):
        i = row_idx[ii]
        cols = [(col_label, col) for col_label, col in row[1].items() if col]
        col_idx = self.columns.get_indexer([x[0] for x in cols])
        for jj, itm in enumerate(cols):
            j = col_idx[jj]
            for pair in itm[1].rstrip(";").split(";"):
                self.ctx[(i, j)].append(pair)


Styler._update_ctx = _update_ctx

bd = Side(style="thin", color="000000")
thin_border = Border(left=bd, top=bd, right=bd, bottom=bd)


def cell_week_summary_table(worksheet, df, startrow):
    percent_format = False
    for col in worksheet.iter_cols(min_row=startrow):
        if percent_format:
            percent_format = False
        else:
            percent_format = True
        for cell in col:
            # column_number = column_index_from_string(cell.column)
            cell.border = thin_border
            column_number = cell.column
            if column_number > (df.columns.get_loc(('Item', 'Part', 'Goal')) + 2):
                if percent_format:
                    cell.number_format = '0.0%'
                worksheet.column_dimensions[cell.column_letter].width = 8


def cell_week_content(worksheet, df, startrow):
    for col in worksheet.iter_cols(min_row=startrow):
        for cell in col:
            # column_number = column_index_from_string(cell.column)
            column_number = cell.column
            if column_number > (df.columns.get_loc(('SUMMARY', 'Yield')) + 1):
                cell.number_format = '0.00%'
                worksheet.column_dimensions[cell.column_letter].width = 10
            elif column_number == 3:  # shipping_date
                worksheet.column_dimensions[cell.column_letter].width = 18.7
            else:  # 0.00% number format
                worksheet.column_dimensions[cell.column_letter].width = 6


def wafer_content(worksheet, df, startrow, size):
    w, h = size
    for row in worksheet.iter_cols(min_row=startrow):
        for cell in row:
            worksheet.column_dimensions[cell.column_letter].width = w
            worksheet.row_dimensions[cell.row].height = h * 5  # try to make perfect circle wafer


def cell_lot_history_table(worksheet, df, startrow):
    # 1 InstNo, 2 Stage,3 Equipment, 4 LotQty, 5 TrackInTime, 6 TrackOutTime, 7 Procedure
    # 8 Step Description, 9 Recipe, 10 TrackInUser, 11 TrackOutUser, 12 LOTID, 0 Events
    cell_width_dict = {1: 8, 2: 13, 3: 9.5, 4: 6, 5: 18.5, 6: 18.5, 7: 12, 8: 34, 9: 14, 10: 13, 11: 13, 12: 10.3,
                       0: 13}
    for col in worksheet.iter_cols(min_row=startrow):
        for cell in col:
            column_number = cell.column
            cell.border = thin_border
            if column_number == 1:  # Item
                worksheet.column_dimensions[cell.column_letter].width = 20
            else:
                i = (column_number - 1) % 13
                worksheet.column_dimensions[cell.column_letter].width = cell_width_dict[i]
                if i in {5, 6}:
                    cell.number_format = 'yyyy-mm-dd hh:mm:ss'


def neworder(sheet_list, to_position=0):
    """Takes a list of ints, and inserts the last int, to tpos location (0-index)"""
    lst = []
    lpos = (len(sheet_list) - 1)
    # Just a counter
    for x in range(len(sheet_list)):
        if x > (to_position - 1) and x != to_position:
            lst.append(x - 1)
        elif x == to_position:
            lst.append(lpos)
        else:
            lst.append(x)

    return lst


def df_update_xlsx(df, filename, sheetname='Sheet1', styler=None, applymap=False, apply_axis=None, offset=2,
                   startrow=None, truncate_sheet=True, multindex_headers=0, header=True, fit_col_width=True,
                   cell_width_height=None, cell_width_height_kwargs=None, apply_kwargs=None, header_style=None,
                   return_writer=False, freeze=None, read_header_style_row=0, mapping_style_dict=None,
                   new_sheet_position=None, pcs_group_idx_optimize=False, xlsm_template=None, **to_excel_kwargs):
    """
    Dump the DataFrame to excel in .xlsx format
    :param df:input DataFrame
    :param filename:File path or existing ExcelWriter(engine = openpyxl) (Example: '/path/to/file.xlsx')
    :param sheetname:Name of sheet which will contain DataFrame. (default: 'Sheet1')
    :param styler: styler object to format output excel
    :param applymap: use 'applymap' to style DataFrame, default using 'apply' method to style
    :param apply_axis: int, str or None,  if use 'apply' method to style, apply to each column (axis=0 or 'index')
           or to each row (axis=1 or 'columns'), or to the entire DataFrame at once with axis=None, default None
    :param offset: offset for column width adjustment, varied with index property
    :param startrow:upper left cell row to dump DataFrame.
                 Per default (startrow=None) calculate the last row
                 in the existing sheet and write to the next row...
    :param truncate_sheet:truncate (remove and recreate) [sheet_name] (refresh)
                       before writing DataFrame to Excel file, default True
    :param multindex_headers: int, input the layer number of multindex headers, default 0
    :param header: bool or list of str, default True. Write out the column names.
                   If a list of string is given it is assumed to be aliases for the column names.
    :param fit_col_width: auto adjust the column width by the max len string in the column, default True
    :param cell_width_height: function to set the cell width and height
    :param cell_width_height_kwargs: kwargs pass cell_width_height func
    :param apply_kwargs: dict, keyword arguments to pass to styler applying _func in addition to the array/series.
    :param header_style: sub_class of NamedStyle, use to style the header, default None
    :param return_writer: determine method to return None or return excel writer object
    :param freeze: default None, the given str are cell like 'A2'.. etc,
                freeze rows above the given cell and columns to the left.
    :param read_header_style_row: default 0, determine the which row to read the key of header style dict
    :param mapping_style_dict: default None, mapping the read_style value to real/existed style name
    :param new_sheet_position: default None, at the end, input zero-indexed number to set new sheet position
    :param to_excel_kwargs:key word arguments which will be passed to 'DataFrame.to_excel()'
    :param pcs_group_idx_optimize: optimize pcs raw_group idx
    :param xlsm_template: template xlsm file to write data and save to other site
    :return: None or writer object
    """

    if apply_kwargs is None:
        apply_kwargs = dict()
    if cell_width_height_kwargs is None:
        cell_width_height_kwargs = dict()
    if 'engine' in to_excel_kwargs:  # ignore [engine] parameter if it was passed
        to_excel_kwargs.pop('engine')

    # writer = StyleFrame.ExcelWriter(file)
    # sf = StyleFrame(df)
    # sf.to_excel(writer, sheet_name='DEFGROUP', best_fit = sf.columns.values.tolist())
    if isinstance(filename, str):
        if not filename.endswith('.xlsx'):
            if not filename.endswith('.xlsm') and xlsm_template is None:
                filename = '{}.xlsx'.format(filename)
        writer = pd.ExcelWriter(filename, engine='openpyxl')

        input_writer = False
    else:
        writer = filename
        input_writer = True

    try:
        if input_writer:
            if not writer.book.sheetnames:
                raise FileNotFoundError
        elif xlsm_template is not None:
            writer.book = load_workbook(xlsm_template, keep_vba=True)
            sheet = writer.book['Template']
            sheet.title = sheetname
        else:
            writer.book = load_workbook(filename)  # try to open an existing workbook
        if startrow is None:  # get the last row in the existing Excel sheet if it was not specified
            if sheetname in writer.book.sheetnames:
                startrow = writer.book[sheetname].max_row
            elif header and not multindex_headers and truncate_sheet:
                startrow = 0
            # elif header and not multindex_headers:
            #     startrow = 0
            else:  # sheet not in excel, will update it
                startrow = 0
        # truncate sheet
        if truncate_sheet and sheetname in writer.book.sheetnames:
            startrow = 0
            idx = writer.book.sheetnames.index(sheetname)  # index of [sheet_name] sheet
            writer.book.remove(writer.book.worksheets[idx])  # remove [sheet_name]
            writer.book.create_sheet(sheetname, idx)  # create an empty sheet [sheet_name] using old index
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}  # copy existing sheets
    except FileNotFoundError:
        if startrow is None:
            startrow = 0
        pass  # file does not exist yet, will create it

    # optimize pcs raw_group idx
    if pcs_group_idx_optimize and '_raw_' in sheetname and startrow > 0:
        pcs_index = startrow - 2
        new_indexes = range(pcs_index, pcs_index + len(df))
        df.index = new_indexes
        if apply_kwargs:
            df_ref = apply_kwargs['df_ref']
            if not df_ref.empty:
                df_ref.index = new_indexes

    if styler is None:
        df.to_excel(writer, sheet_name=sheetname, startrow=startrow, header=header,
                    **to_excel_kwargs)  # send df to writer
    else:
        if applymap:
            styled = df.style.applymap(styler)
        else:
            styled = df.style.apply(styler, **apply_kwargs, axis=apply_axis)
        # if table_styles:
        #         styled.set_table_styles(table_styles=table_styles)
        styled.to_excel(writer, sheet_name=sheetname, startrow=startrow, header=header,
                        **to_excel_kwargs)  # send df.style to writer
    worksheet = writer.sheets[sheetname]  # pull worksheet object

    # try to adjust column width
    if fit_col_width:
        for idx, col in enumerate(df):  # loop through all columns
            series = df[col]
            max_len = max((
                series.astype(str).map(len).max(),  # len of largest item
                len(str(series.name))  # len of column name/header
            )) + 1  # adding a little extra space
            worksheet.column_dimensions[get_column_letter(idx + offset + 1)].width = max_len  # set column width

    # try to fix the bug that pd.to_excel would generate blank row below Multindex header
    _idx = startrow
    if multindex_headers:
        if header:
            _idx = startrow + multindex_headers + 1
        else:
            _idx = startrow + 1
        worksheet.delete_rows(_idx, 1)
    if header and not multindex_headers:
        _idx = startrow + 1
    # _idx: data start writing row index
    # set the number format and column width for weekly data
    if cell_width_height:
        cell_width_height(worksheet, df, _idx, **cell_width_height_kwargs)

    # try to style the Multi-Index header by its contents
    if header and header_style and truncate_sheet:
        n = multindex_headers
        if not n:
            n = 0
        style = str()
        for col in worksheet.iter_cols(min_row=_idx - n, max_row=_idx - 1):  # min_row and max_row: headers range
            read_sty = col[read_header_style_row].value
            if read_sty:
                if mapping_style_dict:
                    style = mapping_style_dict.get(str(read_sty), 'WAFER')
                else:
                    style = read_sty
            elif col[1].value in {'FSV', 'BSV', 'CP'}:  # these elif can be remove in other uses of df_update_excel
                style = col[1].value  # may update to have better compatibility in the future
            for cell in col:
                if style in writer.book.named_styles:
                    cell.style = style
                else:
                    cell.style = header_style(style)
    if freeze and isinstance(freeze, str):
        worksheet.freeze_panes = freeze

    if new_sheet_position is not None:
        wb = writer.book
        currentorder = wb.sheetnames
        myorder = neworder(currentorder, new_sheet_position)
        wb._sheets = [wb._sheets[i] for i in myorder]

    writer.book.properties.creator = '陳天霖 (Ernie Chen)'

    if return_writer:
        return writer
    else:
        try:
            writer.save()
        except PermissionError as err:
            path = writer.path
            new_path = path[:-5] + '-' + path[-5:]
            writer.path = new_path
            writer.save()
            print(str(err)+f"\nfile has been saved to {new_path}, please change back to origin file name.")


def main():
    # df_ooc_group_update()

    # df_group_update(modify=True)

    # df_group_update(modify=False)
    pass


if __name__ == '__main__':
    main()
