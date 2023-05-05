#!/usr/bin/env python
# _*_ coding:utf-8 _*_


import numpy as np
# import pandas as pd
# from openpyxl.utils import get_column_letter
from openpyxl import load_workbook, Workbook


# create excel type item
# wb = Workbook()
# # select the active worksheet
# ws = wb.active
#
# col_counter = 0
# counter = 0
# for column in range(9):  # 1~9
#     column
#     column_letter = get_column_letter(column)
#     for row in range(1, 11):
#         counter = counter + 1
#         ws[column_letter + str(row)] = counter
#
# wb.save("sample.xlsx")


# def formula_gen():
#     formula = '= AVERAGE({})'
#     # s = """=VLOOKUP($M113&"_1",$A$2:$CW$96,MATCH($N113&"_1",$A$1:$CW$1,0),0)"""
#     # s = """=VLOOKUP($M{}&"_{}",$A$2:$CW$96,MATCH($N{}&"_{}",$A$1:$CW$1,0),0)"""
#     for r in range(58):
#         n = r + 113
#         l = list()
#         for i in range(11):
#             m = i + 1
#             s1 = f"""VLOOKUP($M{n}&"_{m}",$A$2:$CW$100,MATCH($N{n}&$O112,$A$1:$CW$1,0),0)"""
#             l.append(s1)
#         result = ", ".join(l)
#         print(formula.format(result))
# def formula_gen():
#     # s = """=AVERAGE({}:{})"""
#     origin_x, origin_y = (6, 2)
#     for col in range(9):
#         for row in range(9):
#             start_x = origin_x + col * 11
#             start_y = origin_y + row * 11
#             end_x = start_x + 10
#             end_y = start_y + 10
#             tag = f"_{row + 1}_{col + 1}"
#             result = f"STDEV.S({get_column_letter(start_x)}{start_y}:{get_column_letter(end_x)}{end_y})"
#             print(f"""{tag}   =IFERROR({result},"")""")
#
#
# def formula_gen2():
#     # s = """=AVERAGE({}:{})"""
#     origin_x, origin_y = (6, 2)
#     for col in range(9):
#         for row in range(9):
#             start_x = origin_x + col * 11
#             start_y = origin_y + row * 11
#             end_x = start_x + 10
#             end_y = start_y + 10
#             tag = f"_{row + 1}_{col + 1}"
#             lst = []
#             for _x in range(start_x, end_x + 1):
#                 for _y in range(start_y, end_y + 1):
#                     your_calc = f'{get_column_letter(_x)}{_y}'
#                     s = f"""=IFERROR(IF({your_calc}=0,"",{your_calc}),"")"""
#                     lst.append(s)
#             result = f"{' '.join(lst)}"
#             print(f"""{tag} {result}""")


def copy_worksheet():
    wb2 = Workbook()
    ws2 = wb2.create_sheet('AP197 Rin boxplot')
    # opening the destination excel file
    # wb2 = xl.load_workbook(dest_xl)
    col_counter = 2
    for idx, i in enumerate(range(26)):
        sub_file = r'C:\Users\ernie.chen\Desktop\ReportGen ver2\AP197 Rin\{}AP197.xlsx'.format(i)
        wb1 = load_workbook(sub_file)
        # opening the source excel file
        sheet_names1 = wb1.sheetnames

        for ws in sheet_names1:
            if "DATA" in str(ws):
                print(f'savaing data of {str(ws)}')
                wafer_id = str(ws).rstrip('DATA')
                index = sheet_names1.index(ws)
                ws1 = wb1.worksheets[index]
                # opening the destination excel file
                # wb2 = xl.load_workbook(dest_xl)

                # ws2 = wb2.create_sheet(str(ws))

                # calculate total number of rows and
                # columns in source excel file
                mr = ws1.max_row
                mc = ws1.max_column

                # copying the cell values from source
                # excel file to destination excel file

                # col
                for j in range(4, mc + 1):
                    lst = []
                    # row
                    for k in range(2, mr + 1):
                        # reading cell value from source excel file
                        c = ws1.cell(row=k, column=j)
                        lst.append(c.value)
                    arr = np.array(lst)
                    q1, q2, q3 = np.percentile(arr, [25, 50, 75])
                    iqr = q3 - q1
                    q2_q1 = q2 - q1
                    q3_q2 = q3 - q2
                    min_of_arr = arr.min()
                    max_of_arr = arr.max()
                    q1_iqr = q1 - 1.5 * iqr
                    q3_iqr = q3 + 1.5 * iqr

                    lower_whisker = max([q1_iqr, min_of_arr])
                    upper_whisker = min([q3_iqr, max_of_arr])
                    w_upper_q3 = upper_whisker - q3
                    q1_w_lower = q1 - lower_whisker

                    min_arr = arr[arr < q1_iqr]

                    if min_arr.size:
                        min_outlier = min_of_arr
                    else:
                        min_outlier = ''

                    max_arr = arr[arr > q3_iqr]
                    if max_arr.size:
                        max_outlier = max_of_arr
                    else:
                        max_outlier = ''
                    if not idx:
                        for r, n in enumerate(['wafer id', 'Q1', 'Q2-Q1', 'Q3-Q2', 'Min outlier', 'Max outlier', 'Wupper-Q3', 'Q1-Wlower']):
                            ws2.cell(row=r + 1, column=1).value = n
                    for r, n in enumerate([wafer_id, q1, q2_q1, q3_q2, min_outlier, max_outlier, w_upper_q3, q1_w_lower]):
                        ws2.cell(row=r + 1, column=col_counter).value = n
                col_counter += 1
                # writing the read value to destination excel file
                # ws2.cell(row=k, column=j-3).value = c.value

    # saving the destination excel file
    wb2.save(r'C:\Users\ernie.chen\Desktop\ReportGen ver2\AP197 Rin\AP197Rin history2.xlsx')


def main():
    copy_worksheet()


if __name__ == '__main__':
    # formula_gen()
    # formula_gen2()
    main()
