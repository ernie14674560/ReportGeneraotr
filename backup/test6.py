#!/usr/bin/env python
# _*_ coding:utf-8 _*_
from openpyxl.utils import get_column_letter


# create excel type item
# wb = Workbook()
# # select the active worksheet
# ws = wb.active
#
# col_counter = 0
# counter = 0
# for column in range(9):  # 1~9
#     column
#     column_letter = get_column_letter(column)
#     for row in range(1, 11):
#         counter = counter + 1
#         ws[column_letter + str(row)] = counter
#
# wb.save("sample.xlsx")


# def formula_gen():
#     formula = '= AVERAGE({})'
#     # s = """=VLOOKUP($M113&"_1",$A$2:$CW$96,MATCH($N113&"_1",$A$1:$CW$1,0),0)"""
#     # s = """=VLOOKUP($M{}&"_{}",$A$2:$CW$96,MATCH($N{}&"_{}",$A$1:$CW$1,0),0)"""
#     for r in range(58):
#         n = r + 113
#         l = list()
#         for i in range(11):
#             m = i + 1
#             s1 = f"""VLOOKUP($M{n}&"_{m}",$A$2:$CW$100,MATCH($N{n}&$O112,$A$1:$CW$1,0),0)"""
#             l.append(s1)
#         result = ", ".join(l)
#         print(formula.format(result))
def formula_gen():
    # s = """=AVERAGE({}:{})"""
    origin_x, origin_y = (6, 2)
    for col in range(9):
        for row in range(9):
            start_x = origin_x + col * 11
            start_y = origin_y + row * 11
            end_x = start_x + 10
            end_y = start_y + 10
            tag = f"_{row + 1}_{col + 1}"
            result = f"STDEV.S({get_column_letter(start_x)}{start_y}:{get_column_letter(end_x)}{end_y})"
            print(f"""{tag}   =IFERROR({result},"")""")


def formula_gen2():
    # s = """=AVERAGE({}:{})"""
    origin_x, origin_y = (6, 2)
    for col in range(9):
        for row in range(9):
            start_x = origin_x + col * 11
            start_y = origin_y + row * 11
            end_x = start_x + 10
            end_y = start_y + 10
            tag = f"_{row + 1}_{col + 1}"
            lst = []
            for _x in range(start_x, end_x+1):
                for _y in range(start_y, end_y+1):
                    your_calc = f'{get_column_letter(_x)}{_y}'
                    s = f"""=IFERROR(IF({your_calc}=0,"",{your_calc}),"")"""
                    lst.append(s)
            result = f"{' '.join(lst)}"
            print(f"""{tag} {result}""")


if __name__ == '__main__':
    # formula_gen()
    formula_gen2()
